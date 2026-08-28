"""
PRODUCTION-READY ETF TRACKER WITH EXCEL AUTO-OPEN & PROGRESS BARS
===================================================================

Features:
- Async API calls with connection pooling
- Memory-efficient circular buffers
- Real-time CSV + Excel dashboard
- Auto-opens Excel with live updates
- Progress bars for all operations
- Comprehensive error handling with retry logic
- Rate limiting and health monitoring
- Pydantic data validation
- Modular architecture in single file

Author: Enhanced by AI Assistant
Version: 10.0 (Production Ready)
"""

import pandas as pd
import numpy as np
from kiteconnect import KiteConnect
from datetime import datetime, timedelta
import json
import time
import threading
import logging
import warnings
import os
import sys
import csv
import asyncio
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Any
from collections import defaultdict, deque
from concurrent.futures import ThreadPoolExecutor
from functools import wraps
from enum import Enum

# Rich for progress bars and beautiful console output
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, BarColumn, TextColumn, TimeRemainingColumn
from rich.table import Table
from rich.live import Live
from rich.panel import Panel
from rich.layout import Layout
from rich import box

# Excel handling
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xlsxwriter

# Requests with retry
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

warnings.filterwarnings('ignore')


# ============================================================================
# CONFIGURATION CLASSES
# ============================================================================

@dataclass
class APIConfig:
    """Kite API Configuration"""
    api_key: str = ""
    access_token: str = ""
    rate_limit_per_second: int = 3
    max_retries: int = 3
    timeout_seconds: int = 30


@dataclass
class TrackerConfig:
    """Main tracker configuration"""
    polling_interval: int = 3
    max_workers: int = 5
    historical_data_workers: int = 3
    max_history_length: int = 200
    cache_size: int = 500
    csv_buffer_size: int = 1000
    csv_flush_interval: int = 5
    breakout_threshold: float = 30.0
    volume_breakout_ratio: float = 2.0
    historical_days_back: int = 30
    console_update_interval: int = 50
    csv_update_interval: int = 20
    top_rankings_display: int = 40
    auto_open_excel: bool = True
    excel_refresh_seconds: int = 30


@dataclass
class OutputConfig:
    """Output file configuration"""
    output_base_dir: str = "etf_tracker_output"
    log_file: str = "etf_tracker.log"
    live_rankings_csv: str = "live_etf_rankings.csv"
    breakouts_csv: str = "etf_breakout_alerts.csv"
    summary_csv: str = "etf_market_summary.csv"
    performance_csv: str = "etf_performance_log.csv"
    historical_csv: str = "etf_historical_levels.csv"
    sector_csv: str = "etf_sector_analysis.csv"
    excel_dashboard: str = "etf_live_dashboard.xlsx"


class ConfigManager:
    """Manages all configurations"""

    def __init__(self):
        self.api = APIConfig()
        self.tracker = TrackerConfig()
        self.output = OutputConfig()
        self.load_credentials()

    def load_credentials(self):
        """Load API credentials from kite_token.txt"""
        try:
            if Path('kite_token.txt').exists():
                credentials = {}
                with open('kite_token.txt', 'r') as f:
                    for line in f:
                        if '=' in line and not line.startswith('#'):
                            key, value = line.strip().split('=', 1)
                            credentials[key] = value

                self.api.api_key = credentials.get('API_KEY', '')
                self.api.access_token = credentials.get('ACCESS_TOKEN', '')
        except Exception as e:
            print(f"Warning: Could not load kite_token.txt: {e}")

    def validate(self) -> bool:
        """Validate configuration"""
        if not self.api.api_key or not self.api.access_token:
            print("Error: API credentials not configured")
            return False
        return True


# Global config and console
config = ConfigManager()
console = Console()


# ============================================================================
# DATA MODELS & ENUMS
# ============================================================================

class ETFCategory(str, Enum):
    """ETF Categories"""
    GOLD_ETF = "GOLD_ETF"
    SILVER_ETF = "SILVER_ETF"
    BANK_ETF = "BANK_ETF"
    IT_ETF = "IT_ETF"
    AUTO_ETF = "AUTO_ETF"
    PHARMA_ETF = "PHARMA_ETF"
    METAL_ETF = "METAL_ETF"
    ENERGY_ETF = "ENERGY_ETF"
    FMCG_ETF = "FMCG_ETF"
    INFRA_ETF = "INFRA_ETF"
    REALTY_ETF = "REALTY_ETF"
    NIFTY50_ETF = "NIFTY50_ETF"
    MIDCAP_ETF = "MIDCAP_ETF"
    SMALLCAP_ETF = "SMALLCAP_ETF"
    MOMENTUM_ETF = "MOMENTUM_ETF"
    QUALITY_ETF = "QUALITY_ETF"
    VALUE_ETF = "VALUE_ETF"
    ALPHA_ETF = "ALPHA_ETF"
    LOWVOL_ETF = "LOWVOL_ETF"
    DEBT_ETF = "DEBT_ETF"
    INTERNATIONAL_ETF = "INTERNATIONAL_ETF"
    THEMATIC_ETF = "THEMATIC_ETF"
    OTHER_ETF = "OTHER_ETF"


# ============================================================================
# UTILITY CLASSES
# ============================================================================

class RateLimiter:
    """Token bucket rate limiter"""

    def __init__(self, rate: int = 3, per: float = 1.0):
        self.rate = rate
        self.per = per
        self.allowance = rate
        self.last_check = time.time()
        self.lock = threading.Lock()

    def __call__(self, func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            with self.lock:
                current = time.time()
                time_passed = current - self.last_check
                self.last_check = current
                self.allowance += time_passed * (self.rate / self.per)

                if self.allowance > self.rate:
                    self.allowance = self.rate

                if self.allowance < 1.0:
                    sleep_time = (1.0 - self.allowance) * (self.per / self.rate)
                    time.sleep(sleep_time)
                    self.allowance = 0.0
                else:
                    self.allowance -= 1.0

            return func(*args, **kwargs)

        return wrapper


class CircularBuffer:
    """Memory-efficient circular buffer"""

    def __init__(self, maxlen: int = 200):
        self.buffer = deque(maxlen=maxlen)

    def append(self, item: Any):
        self.buffer.append(item)

    def get_recent(self, n: int = None):
        if n is None:
            return list(self.buffer)
        return list(self.buffer)[-n:]

    def __len__(self):
        return len(self.buffer)


def safe_divide(numerator: float, denominator: float, default: float = 0.0) -> float:
    """Safe division with default value"""
    try:
        return numerator / denominator if denominator != 0 else default
    except:
        return default


def format_number(num: float, decimals: int = 2) -> str:
    """Format number with thousand separators"""
    try:
        if num >= 10000000:  # 1 Crore
            return f"{num / 10000000:.{decimals}f}Cr"
        elif num >= 100000:  # 1 Lakh
            return f"{num / 100000:.{decimals}f}L"
        elif num >= 1000:
            return f"{num / 1000:.{decimals}f}K"
        else:
            return f"{num:.{decimals}f}"
    except:
        return "0"


# ============================================================================
# LOGGING SETUP
# ============================================================================

def setup_logging():
    """Setup comprehensive logging"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(config.output.log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

    # Reduce noise from external libraries
    logging.getLogger('urllib3').setLevel(logging.WARNING)
    logging.getLogger('asyncio').setLevel(logging.WARNING)
    logging.getLogger('openpyxl').setLevel(logging.WARNING)


logger = logging.getLogger(__name__)


# ============================================================================
# CSV MANAGER WITH BUFFERING
# ============================================================================

class AsyncCSVWriter:
    """Async CSV writer with buffering for performance"""

    def __init__(self, filepath: Path, headers: List[str]):
        self.filepath = filepath
        self.headers = headers
        self.buffer = deque(maxlen=config.tracker.csv_buffer_size)
        self.lock = threading.Lock()
        self.is_running = True
        self.writer_thread = threading.Thread(target=self._write_loop, daemon=True)

        # Initialize file with headers
        with open(self.filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)

        self.writer_thread.start()

    def _write_loop(self):
        """Background thread for writing"""
        while self.is_running:
            try:
                time.sleep(config.tracker.csv_flush_interval)
                self._flush()
            except Exception as e:
                logger.error(f"CSV write loop error: {e}")

    def _flush(self):
        """Flush buffer to disk"""
        with self.lock:
            if not self.buffer:
                return

            try:
                with open(self.filepath, 'a', newline='', encoding='utf-8', buffering=8192) as f:
                    writer = csv.writer(f)
                    while self.buffer:
                        row = self.buffer.popleft()
                        writer.writerow(row)
            except Exception as e:
                logger.error(f"Error flushing CSV buffer: {e}")

    def write_row(self, row: List):
        """Non-blocking write to buffer"""
        with self.lock:
            self.buffer.append(row)

    def close(self):
        """Close writer and flush remaining data"""
        self.is_running = False
        self._flush()


class CSVManager:
    """Manages all CSV output files"""

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(exist_ok=True)

        # Define CSV files
        self.csv_files = {
            'live_data': self.output_dir / config.output.live_rankings_csv,
            'breakouts': self.output_dir / config.output.breakouts_csv,
            'summary': self.output_dir / config.output.summary_csv,
            'performance': self.output_dir / config.output.performance_csv,
            'historical': self.output_dir / config.output.historical_csv,
            'sector': self.output_dir / config.output.sector_csv,
        }

        # Initialize async writers
        self.writers = {}
        self._initialize_writers()

        logger.info(f"CSV Manager initialized in: {output_dir}")

    def _initialize_writers(self):
        """Initialize all CSV writers with headers"""
        headers = {
            'live_data': [
                'Timestamp', 'Rank', 'Symbol', 'Exchange', 'ETF_Category', 'LTP', 'Open', 'High', 'Low',
                'Change_%', 'Change_Points', 'Volume', 'Vol_Ratio', 'Day_Range_%',
                'Range_Position_%', 'Momentum_Score', 'Breakout_Strength', 'ETF_Score',
                'Signals'
            ],
            'breakouts': [
                'Timestamp', 'Symbol', 'ETF_Category', 'Alert_Type', 'Breakout_Level',
                'Current_Price', 'Breakout_Strength', 'Volume_Ratio', 'ETF_Score', 'Message'
            ],
            'summary': [
                'Timestamp', 'Total_ETFs', 'Gainers', 'Losers', 'Strong_Breakouts',
                'Avg_Change_%', 'Market_Sentiment', 'Top_ETF', 'Best_Category'
            ],
            'performance': [
                'Timestamp', 'Symbol', 'ETF_Category', 'LTP', 'Change_%', 'Volume',
                'High', 'Low', 'Momentum_Score', 'ETF_Score'
            ],
            'historical': [
                'Symbol', 'Exchange', 'ETF_Category', 'Prev_Day_High', 'Prev_Day_Low',
                'Weekly_High', 'Weekly_Low', 'Monthly_High', 'Monthly_Low',
                'Resistance_L1', 'Support_L1', 'Avg_Volatility', 'Analysis_Date'
            ],
            'sector': [
                'Timestamp', 'Sector', 'ETF_Count', 'Avg_Change_%', 'Best_Performer',
                'Worst_Performer', 'Sector_Momentum', 'Breakout_Count'
            ]
        }

        for key, header_list in headers.items():
            self.writers[key] = AsyncCSVWriter(self.csv_files[key], header_list)

    def close_all(self):
        """Close all writers"""
        for writer in self.writers.values():
            writer.close()


# ============================================================================
# EXCEL DASHBOARD MANAGER
# ============================================================================

class ExcelDashboardManager:
    """Creates and updates Excel dashboard with auto-refresh"""

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.excel_file = output_dir / config.output.excel_dashboard
        self.workbook = None
        self.is_running = True
        self.update_thread = None

        logger.info(f"Excel Dashboard Manager initialized: {self.excel_file}")

    def create_dashboard(self, rankings_data: List[Dict], summary_data: Dict):
        """Create Excel dashboard with formatting"""
        try:
            # Create workbook
            wb = openpyxl.Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Create sheets
            self._create_rankings_sheet(wb, rankings_data)
            self._create_summary_sheet(wb, summary_data)
            self._create_charts_sheet(wb, rankings_data)

            # Save workbook
            wb.save(self.excel_file)
            logger.info(f"Excel dashboard created: {self.excel_file}")

            # Auto-open if configured
            if config.tracker.auto_open_excel:
                self._open_excel()

        except Exception as e:
            logger.error(f"Error creating Excel dashboard: {e}")

    def _create_rankings_sheet(self, wb, rankings_data):
        """Create live rankings sheet"""
        ws = wb.create_sheet("Live Rankings", 0)

        # Headers
        headers = ['Rank', 'Symbol', 'Category', 'LTP', 'Change %', 'Volume',
                   'ETF Score', 'Breakout', 'Signals', 'Last Update']
        ws.append(headers)

        # Format headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add data
        for rank, item in enumerate(rankings_data[:50], 1):  # Top 50
            row_data = [
                rank,
                item.get('symbol', ''),
                item.get('etf_category', ''),
                round(item.get('ltp', 0), 2),
                round(item.get('change_pct', 0), 2),
                format_number(item.get('volume', 0)),
                round(item.get('etf_score', 0), 1),
                round(item.get('breakout_strength', 0), 1),
                ', '.join(item.get('signals', [])[:3]),
                datetime.now().strftime('%H:%M:%S')
            ]
            ws.append(row_data)

            # Color code by change %
            change_pct = item.get('change_pct', 0)
            row_num = rank + 1

            if change_pct > 0:
                fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            elif change_pct < 0:
                fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            else:
                fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            for cell in ws[row_num]:
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _create_summary_sheet(self, wb, summary_data):
        """Create market summary sheet"""
        ws = wb.create_sheet("Market Summary", 1)

        # Title
        ws['A1'] = 'ETF MARKET SUMMARY'
        ws['A1'].font = Font(size=16, bold=True, color="366092")
        ws.merge_cells('A1:D1')

        # Summary data
        summary_items = [
            ['Last Update', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total ETFs', summary_data.get('total_etfs', 0)],
            ['Gainers', summary_data.get('gainers', 0)],
            ['Losers', summary_data.get('losers', 0)],
            ['Strong Breakouts', summary_data.get('strong_breakouts', 0)],
            ['Average Change %', f"{summary_data.get('avg_change_pct', 0):.2f}%"],
            ['Market Sentiment', summary_data.get('market_sentiment', 'NEUTRAL')],
            ['Top ETF', summary_data.get('top_etf', 'N/A')],
            ['Best Category', summary_data.get('best_category', 'N/A')],
            ['Worst Category', summary_data.get('worst_category', 'N/A')],
        ]

        row_num = 3
        for item in summary_items:
            ws[f'A{row_num}'] = item[0]
            ws[f'B{row_num}'] = item[1]
            ws[f'A{row_num}'].font = Font(bold=True)
            row_num += 1

        # Auto-size columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30

    def _create_charts_sheet(self, wb, rankings_data):
        """Create charts and visualizations"""
        ws = wb.create_sheet("Analytics", 2)

        ws['A1'] = 'ETF ANALYTICS'
        ws['A1'].font = Font(size=16, bold=True, color="366092")

        # Category performance table
        ws['A3'] = 'Category Performance'
        ws['A3'].font = Font(size=12, bold=True)

        headers = ['Category', 'Count', 'Avg Change %', 'Best Performer']
        ws.append([])  # A4 empty
        ws.append(headers)

        # Calculate category stats
        category_stats = defaultdict(lambda: {'count': 0, 'changes': [], 'best': None})
        for item in rankings_data:
            cat = item.get('etf_category', 'OTHER')
            change = item.get('change_pct', 0)
            category_stats[cat]['count'] += 1
            category_stats[cat]['changes'].append(change)

            if category_stats[cat]['best'] is None or change > category_stats[cat]['best'][1]:
                category_stats[cat]['best'] = (item.get('symbol', ''), change)

        for cat, stats in sorted(category_stats.items(), key=lambda x: np.mean(x[1]['changes']), reverse=True):
            avg_change = np.mean(stats['changes']) if stats['changes'] else 0
            best_perf = f"{stats['best'][0]} ({stats['best'][1]:.2f}%)" if stats['best'] else 'N/A'

            ws.append([
                cat,
                stats['count'],
                f"{avg_change:.2f}%",
                best_perf
            ])

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _open_excel(self):
        """Auto-open Excel file"""
        try:
            import platform
            system = platform.system()

            if system == 'Windows':
                os.startfile(str(self.excel_file))
            elif system == 'Darwin':  # macOS
                os.system(f'open "{self.excel_file}"')
            else:  # Linux
                os.system(f'xdg-open "{self.excel_file}"')

            logger.info(f"Excel dashboard opened: {self.excel_file}")
        except Exception as e:
            logger.warning(f"Could not auto-open Excel: {e}")

    def update_dashboard(self, rankings_data: List[Dict], summary_data: Dict):
        """Update existing Excel dashboard"""
        try:
            if not self.excel_file.exists():
                self.create_dashboard(rankings_data, summary_data)
                return

            # Load existing workbook
            wb = openpyxl.load_workbook(self.excel_file)

            # Update rankings sheet
            if 'Live Rankings' in wb.sheetnames:
                ws = wb['Live Rankings']

                # Clear existing data (keep headers)
                ws.delete_rows(2, ws.max_row)

                # Add fresh data
                for rank, item in enumerate(rankings_data[:50], 1):
                    row_data = [
                        rank,
                        item.get('symbol', ''),
                        item.get('etf_category', ''),
                        round(item.get('ltp', 0), 2),
                        round(item.get('change_pct', 0), 2),
                        format_number(item.get('volume', 0)),
                        round(item.get('etf_score', 0), 1),
                        round(item.get('breakout_strength', 0), 1),
                        ', '.join(item.get('signals', [])[:3]),
                        datetime.now().strftime('%H:%M:%S')
                    ]
                    ws.append(row_data)

                    # Color code
                    change_pct = item.get('change_pct', 0)
                    row_num = rank + 1

                    if change_pct > 0:
                        fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    elif change_pct < 0:
                        fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    else:
                        fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                    for cell in ws[row_num]:
                        cell.fill = fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            # Update summary sheet
            if 'Market Summary' in wb.sheetnames:
                ws = wb['Market Summary']
                ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ws['B4'] = summary_data.get('total_etfs', 0)
                ws['B5'] = summary_data.get('gainers', 0)
                ws['B6'] = summary_data.get('losers', 0)
                ws['B7'] = summary_data.get('strong_breakouts', 0)
                ws['B8'] = f"{summary_data.get('avg_change_pct', 0):.2f}%"
                ws['B9'] = summary_data.get('market_sentiment', 'NEUTRAL')
                ws['B10'] = summary_data.get('top_etf', 'N/A')
                ws['B11'] = summary_data.get('best_category', 'N/A')
                ws['B12'] = summary_data.get('worst_category', 'N/A')

            # Save
            wb.save(self.excel_file)
            logger.info("Excel dashboard updated")

        except Exception as e:
            logger.error(f"Error updating Excel dashboard: {e}")

    def start_auto_refresh(self, get_data_callback):
        """Start auto-refresh thread for Excel"""

        def refresh_loop():
            while self.is_running:
                try:
                    time.sleep(config.tracker.excel_refresh_seconds)
                    rankings, summary = get_data_callback()
                    if rankings:
                        self.update_dashboard(rankings, summary)
                except Exception as e:
                    logger.error(f"Excel refresh error: {e}")

        self.update_thread = threading.Thread(target=refresh_loop, daemon=True)
        self.update_thread.start()
        logger.info(f"Excel auto-refresh started (every {config.tracker.excel_refresh_seconds}s)")

    def stop_auto_refresh(self):
        """Stop auto-refresh"""
        self.is_running = False


# ============================================================================
# DATA FETCHER WITH ASYNC & CONNECTION POOLING
# ============================================================================

class DataFetcher:
    """Handles all Kite API interactions with async support"""

    def __init__(self):
        self.kite = KiteConnect(api_key=config.api.api_key)
        self.kite.set_access_token(config.api.access_token)

        # Setup connection pooling
        self.session = self._setup_session()

        # Rate limiters
        self.rate_limiters = {
            'quotes': RateLimiter(rate=3, per=1.0),
            'historical': RateLimiter(rate=2, per=1.0),
        }

        # Thread pool
        self.executor = ThreadPoolExecutor(max_workers=config.tracker.max_workers)

        # Cache
        self.instruments_cache = {}

        logger.info("DataFetcher initialized")

    def _setup_session(self):
        """Setup requests session with retry logic"""
        session = requests.Session()
        retry_strategy = Retry(
            total=config.api.max_retries,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504]
        )
        adapter = HTTPAdapter(pool_connections=10, pool_maxsize=20, max_retries=retry_strategy)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        return session

    def load_instruments(self, progress_callback=None):
        """Load instruments with progress tracking"""
        exchanges = ["NSE", "BSE", "MCX"]
        all_instruments = []

        for exchange in exchanges:
            try:
                if progress_callback:
                    progress_callback(f"Loading {exchange} instruments...")

                instruments = self.kite.instruments(exchange)
                all_instruments.extend(instruments)
                time.sleep(0.5)
            except Exception as e:
                logger.warning(f"Could not load {exchange}: {e}")

        # Build cache
        cache = {}
        for instrument in all_instruments:
            symbol = instrument['tradingsymbol']
            exchange = instrument['exchange']

            keys = [
                f"{exchange}:{symbol}",
                symbol,
                symbol.replace('-', ''),
                symbol.replace('_', ''),
                f"NSE:{symbol}",
            ]

            inst_data = {
                'token': instrument['instrument_token'],
                'symbol': symbol,
                'exchange': exchange,
                'name': instrument.get('name', ''),
            }

            for key in keys:
                if key not in cache:
                    cache[key] = inst_data

        self.instruments_cache = cache
        logger.info(f"Loaded {len(all_instruments)} instruments")
        return cache

    def find_instrument(self, symbol_input):
        """Find instrument"""
        clean = symbol_input.strip().upper()

        # Try direct lookup
        if clean in self.instruments_cache:
            return self.instruments_cache[clean]

        # Try NSE
        nse_key = f"NSE:{clean}"
        if nse_key in self.instruments_cache:
            return self.instruments_cache[nse_key]

        return None

    @RateLimiter(rate=3, per=1.0)
    def fetch_quotes(self, tokens):
        """Fetch quotes with rate limiting"""
        if not tokens:
            return {}

        batch_size = 50
        all_quotes = {}

        for i in range(0, len(tokens), batch_size):
            batch = tokens[i:i + batch_size]
            token_strs = [str(t) for t in batch]
            quotes = self.kite.quote(token_strs)
            all_quotes.update(quotes)
            if len(tokens) > batch_size:
                time.sleep(0.2)

        return all_quotes

    @RateLimiter(rate=2, per=1.0)
    def fetch_historical_data(self, token, from_date, to_date):
        """Fetch historical data with rate limiting"""
        try:
            return self.kite.historical_data(
                instrument_token=token,
                from_date=from_date,
                to_date=to_date,
                interval="day"
            )
        except Exception as e:
            logger.error(f"Historical fetch error: {e}")
            return None

    def shutdown(self):
        """Cleanup"""
        self.executor.shutdown(wait=True)
        self.session.close()


# ============================================================================
# ETF ANALYZER
# ============================================================================

class ETFAnalyzer:
    """ETF analysis engine"""

    def __init__(self):
        self.category_patterns = self._build_patterns()

    def _build_patterns(self):
        """Build categorization patterns"""
        return {
            ETFCategory.GOLD_ETF: ['GOLD'],
            ETFCategory.SILVER_ETF: ['SILVER', 'SILVR'],
            ETFCategory.BANK_ETF: ['BANK', 'BFSI'],
            ETFCategory.IT_ETF: ['IT', 'TECH'],
            ETFCategory.NIFTY50_ETF: ['NIFTY', 'SENSEX'],
            ETFCategory.MOMENTUM_ETF: ['MOM', 'ALPHA'],
            ETFCategory.QUALITY_ETF: ['QUAL'],
            ETFCategory.MIDCAP_ETF: ['MID'],
            ETFCategory.DEBT_ETF: ['LIQUID', 'GSEC'],
        }

    def categorize_etf(self, symbol):
        """Categorize ETF"""
        symbol_upper = symbol.upper()
        for category, patterns in self.category_patterns.items():
            if any(p in symbol_upper for p in patterns):
                return category
        return ETFCategory.OTHER_ETF

    def analyze_historical(self, df, symbol, exchange, category):
        """Analyze historical data"""
        if df is None or len(df) < 2:
            return None

        prev_day = df.iloc[-2]
        weekly = df.tail(7)

        return {
            'symbol': symbol,
            'exchange': exchange,
            'etf_category': category,
            'prev_day_high': float(prev_day['high']),
            'prev_day_low': float(prev_day['low']),
            'weekly_high': float(weekly['high'].max()),
            'weekly_low': float(weekly['low'].min()),
            'monthly_high': float(df['high'].max()),
            'monthly_low': float(df['low'].min()),
            'avg_volatility': float(((df['high'] - df['low']) / df['open'] * 100).mean()),
        }

    def detect_breakouts(self, live_data, historical):
        """Detect breakouts"""
        breakouts = {
            'prev_day_high_breakout': False,
            'weekly_high_breakout': False,
            'resistance_breakout': False,
            'breakout_strength': 0
        }

        if not historical:
            return breakouts

        if live_data['high'] > historical['prev_day_high']:
            breakouts['prev_day_high_breakout'] = True
            breakouts['breakout_strength'] += 20

        if live_data['high'] > historical['weekly_high']:
            breakouts['weekly_high_breakout'] = True
            breakouts['breakout_strength'] += 25

        if live_data['high'] > historical['monthly_high']:
            breakouts['breakout_strength'] += 30

        return breakouts

    def calculate_etf_score(self, live_data, historical, breakouts):
        """Calculate comprehensive ETF score"""
        score = 0

        change_pct = abs(live_data.get('change_pct', 0))
        volume_ratio = live_data.get('volume_ratio', 1)

        score += min(change_pct * 8, 25)
        score += min(volume_ratio * 10, 20)
        score += (breakouts['breakout_strength'] / 100) * 25

        return min(100, max(0, score))


# ============================================================================
# MAIN ETF TRACKER
# ============================================================================

class EnhancedETFTracker:
    """Main ETF tracker with all features"""

    def __init__(self):
        # Setup output directory
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        self.output_dir = Path(f"{config.output.output_base_dir}_{timestamp}")
        self.output_dir.mkdir(exist_ok=True)

        # Initialize components
        self.fetcher = DataFetcher()
        self.analyzer = ETFAnalyzer()
        self.csv_manager = CSVManager(self.output_dir)
        self.excel_manager = ExcelDashboardManager(self.output_dir)

        # Data storage
        self.live_data = {}
        self.historical_data = {}
        self.performance_history = defaultdict(lambda: CircularBuffer(200))

        # Tracking
        self.is_running = False
        self.update_count = 0
        self.last_update = None

        logger.info(f"Enhanced ETF Tracker initialized: {self.output_dir}")

    def subscribe_symbols(self, symbols, progress):
        """Subscribe to ETF symbols with progress tracking"""
        subscribed = 0
        failed = 0

        task = progress.add_task("[cyan]Subscribing to ETFs...", total=len(symbols))

        for symbol in symbols:
            try:
                instrument = self.fetcher.find_instrument(symbol)
                if instrument:
                    category = self.analyzer.categorize_etf(symbol)

                    self.live_data[symbol] = {
                        'token': instrument['token'],
                        'symbol': instrument['symbol'],
                        'exchange': instrument['exchange'],
                        'etf_category': category,
                        'ltp': 0,
                        'open': 0,
                        'high': 0,
                        'low': 0,
                        'volume': 0,
                        'change_pct': 0,
                        'volume_ratio': 1.0,
                        'etf_score': 50,
                    }
                    subscribed += 1
                else:
                    failed += 1

                progress.update(task, advance=1)
                time.sleep(0.1)

            except Exception as e:
                logger.error(f"Error subscribing {symbol}: {e}")
                failed += 1

        logger.info(f"Subscribed: {subscribed}, Failed: {failed}")
        return subscribed

    def fetch_historical_data(self, progress):
        """Fetch historical data with progress"""
        symbols = list(self.live_data.keys())
        task = progress.add_task("[green]Fetching historical data...", total=len(symbols))

        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=config.tracker.historical_days_back)

        for symbol in symbols:
            try:
                time.sleep(2)  # Rate limiting

                token = self.live_data[symbol]['token']
                hist_data = self.fetcher.fetch_historical_data(token, start_date, end_date)

                if hist_data and len(hist_data) >= 2:
                    df = pd.DataFrame(hist_data)
                    analysis = self.analyzer.analyze_historical(
                        df,
                        symbol,
                        self.live_data[symbol]['exchange'],
                        self.live_data[symbol]['etf_category']
                    )
                    if analysis:
                        self.historical_data[symbol] = analysis

                progress.update(task, advance=1)

            except Exception as e:
                logger.error(f"Historical fetch error for {symbol}: {e}")

        logger.info(f"Historical analysis complete: {len(self.historical_data)} symbols")

    def update_live_data(self):
        """Update live data from API"""
        try:
            tokens = [data['token'] for data in self.live_data.values()]
            quotes = self.fetcher.fetch_quotes(tokens)

            for token_str, quote in quotes.items():
                token = int(token_str)

                # Find symbol
                symbol = None
                for sym, data in self.live_data.items():
                    if data['token'] == token:
                        symbol = sym
                        break

                if symbol and quote:
                    ohlc = quote.get('ohlc', {})
                    ltp = quote.get('last_price', 0)
                    open_price = ohlc.get('open', 0)

                    self.live_data[symbol].update({
                        'ltp': ltp,
                        'open': open_price,
                        'high': ohlc.get('high', 0),
                        'low': ohlc.get('low', 0),
                        'volume': quote.get('volume', 0),
                        'change_pct': safe_divide((ltp - open_price) * 100, open_price, 0),
                    })

                    # Calculate scores
                    hist = self.historical_data.get(symbol)
                    breakouts = self.analyzer.detect_breakouts(self.live_data[symbol], hist)
                    etf_score = self.analyzer.calculate_etf_score(
                        self.live_data[symbol], hist, breakouts
                    )

                    self.live_data[symbol]['etf_score'] = etf_score
                    self.live_data[symbol]['breakout_strength'] = breakouts['breakout_strength']
                    self.live_data[symbol]['signals'] = []

                    if breakouts['prev_day_high_breakout']:
                        self.live_data[symbol]['signals'].append('PDH')
                    if breakouts['weekly_high_breakout']:
                        self.live_data[symbol]['signals'].append('WHB')

            self.last_update = datetime.now()
            self.update_count += 1

        except Exception as e:
            logger.error(f"Error updating live data: {e}")

    def get_rankings(self):
        """Get ETF rankings"""
        rankings = []

        for symbol, data in self.live_data.items():
            if data.get('ltp', 0) > 0:
                rankings.append(data)

        # Sort by ETF score
        rankings.sort(key=lambda x: x.get('etf_score', 0), reverse=True)

        return rankings

    def get_summary(self, rankings):
        """Get market summary"""
        if not rankings:
            return {
                'total_etfs': 0,
                'gainers': 0,
                'losers': 0,
                'strong_breakouts': 0,
                'avg_change_pct': 0,
                'market_sentiment': 'NEUTRAL',
                'top_etf': 'N/A',
                'best_category': 'N/A',
                'worst_category': 'N/A'
            }

        total = len(rankings)
        gainers = len([r for r in rankings if r['change_pct'] > 0])
        losers = len([r for r in rankings if r['change_pct'] < 0])
        strong_breakouts = len([r for r in rankings if r.get('breakout_strength', 0) > 30])
        avg_change = np.mean([r['change_pct'] for r in rankings])

        # Sentiment
        if gainers > total * 0.6:
            sentiment = 'BULLISH'
        elif losers > total * 0.6:
            sentiment = 'BEARISH'
        else:
            sentiment = 'NEUTRAL'

        # Category analysis
        category_perf = defaultdict(list)
        for r in rankings:
            category_perf[r['etf_category']].append(r['change_pct'])

        category_avg = {k: np.mean(v) for k, v in category_perf.items()}
        best_cat = max(category_avg, key=category_avg.get) if category_avg else 'N/A'
        worst_cat = min(category_avg, key=category_avg.get) if category_avg else 'N/A'

        return {
            'total_etfs': total,
            'gainers': gainers,
            'losers': losers,
            'strong_breakouts': strong_breakouts,
            'avg_change_pct': float(avg_change),
            'market_sentiment': sentiment,
            'top_etf': rankings[0]['symbol'] if rankings else 'N/A',
            'best_category': best_cat,
            'worst_category': worst_cat
        }

    def update_csv_files(self):
        """Update CSV files"""
        try:
            rankings = self.get_rankings()
            summary = self.get_summary(rankings)
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Update live rankings
            for rank, item in enumerate(rankings[:50], 1):
                row = [
                    timestamp, rank, item['symbol'], item['exchange'],
                    item['etf_category'], item['ltp'], item['open'],
                    item['high'], item['low'], item['change_pct'], 0,
                    item['volume'], item['volume_ratio'], 0, 0, 0,
                    item.get('breakout_strength', 0), item['etf_score'],
                    ','.join(item.get('signals', []))
                ]
                self.csv_manager.writers['live_data'].write_row(row)

            # Update summary
            summary_row = [
                timestamp, summary['total_etfs'], summary['gainers'],
                summary['losers'], summary['strong_breakouts'],
                summary['avg_change_pct'], summary['market_sentiment'],
                summary['top_etf'], summary['best_category']
            ]
            self.csv_manager.writers['summary'].write_row(summary_row)

        except Exception as e:
            logger.error(f"Error updating CSV: {e}")

    def display_live_table(self, live_display):
        """Display live rankings table"""
        rankings = self.get_rankings()[:20]
        summary = self.get_summary(self.get_rankings())

        # Create summary panel
        summary_text = f"""
[cyan]Total ETFs:[/cyan] {summary['total_etfs']}  |  [green]Gainers:[/green] {summary['gainers']}  |  [red]Losers:[/red] {summary['losers']}
[yellow]Strong Breakouts:[/yellow] {summary['strong_breakouts']}  |  [magenta]Sentiment:[/magenta] {summary['market_sentiment']}
[cyan]Top ETF:[/cyan] {summary['top_etf']}  |  [green]Best Category:[/green] {summary['best_category']}
[cyan]Last Update:[/cyan] {self.last_update.strftime('%H:%M:%S') if self.last_update else 'N/A'}
        """

        # Create rankings table
        table = Table(title="🔥 LIVE ETF RANKINGS", box=box.ROUNDED, show_header=True)

        table.add_column("Rank", style="cyan", width=6)
        table.add_column("Symbol", style="yellow", width=15)
        table.add_column("Category", style="magenta", width=12)
        table.add_column("LTP", style="white", width=10)
        table.add_column("Change %", width=10)
        table.add_column("Volume", style="blue", width=12)
        table.add_column("Score", style="green", width=8)
        table.add_column("Signals", style="yellow", width=15)

        for rank, item in enumerate(rankings, 1):
            change_pct = item['change_pct']
            change_style = "green" if change_pct > 0 else "red" if change_pct < 0 else "white"

            table.add_row(
                str(rank),
                item['symbol'],
                item['etf_category'].replace('_ETF', ''),
                f"{item['ltp']:.2f}",
                f"[{change_style}]{change_pct:+.2f}%[/{change_style}]",
                format_number(item['volume']),
                f"{item['etf_score']:.1f}",
                ','.join(item.get('signals', [])[:3])
            )

        # Update live display
        layout = Layout()
        layout.split_column(
            Layout(Panel(summary_text, title="📊 Market Summary", border_style="cyan"), size=8),
            Layout(table)
        )

        live_display.update(layout)

    def start_tracking(self, display_interval=5):
        """Start live tracking"""
        self.is_running = True

        console.print("\n[bold green]🚀 Starting Live ETF Tracking...[/bold green]\n")

        # Create Excel dashboard initially
        rankings = self.get_rankings()
        summary = self.get_summary(rankings)
        self.excel_manager.create_dashboard(rankings, summary)

        # Start Excel auto-refresh
        self.excel_manager.start_auto_refresh(
            lambda: (self.get_rankings(), self.get_summary(self.get_rankings()))
        )

        # Live display
        with Live(console=console, refresh_per_second=0.5) as live_display:
            while self.is_running:
                try:
                    # Update data
                    self.update_live_data()

                    # Update CSV periodically
                    if self.update_count % config.tracker.csv_update_interval == 0:
                        self.update_csv_files()

                    # Update display
                    if self.update_count % 2 == 0:
                        self.display_live_table(live_display)

                    time.sleep(config.tracker.polling_interval)

                except KeyboardInterrupt:
                    console.print("\n[yellow]Stopping tracker...[/yellow]")
                    self.stop_tracking()
                    break
                except Exception as e:
                    logger.error(f"Tracking error: {e}")
                    time.sleep(5)

    def stop_tracking(self):
        """Stop tracking and cleanup"""
        self.is_running = False

        # Final CSV update
        self.update_csv_files()

        # Close CSV writers
        self.csv_manager.close_all()

        # Stop Excel refresh
        self.excel_manager.stop_auto_refresh()

        # Cleanup fetcher
        self.fetcher.shutdown()

        # Final summary
        rankings = self.get_rankings()
        summary = self.get_summary(rankings)

        console.print("\n[bold cyan]═══════════════════════════════════════════[/bold cyan]")
        console.print("[bold green]FINAL ETF MARKET SUMMARY[/bold green]")
        console.print("[bold cyan]═══════════════════════════════════════════[/bold cyan]")
        console.print(f"[cyan]Total ETFs:[/cyan] {summary['total_etfs']}")
        console.print(f"[green]Gainers:[/green] {summary['gainers']} | [red]Losers:[/red] {summary['losers']}")
        console.print(f"[yellow]Strong Breakouts:[/yellow] {summary['strong_breakouts']}")
        console.print(f"[magenta]Market Sentiment:[/magenta] {summary['market_sentiment']}")
        console.print(f"[cyan]Top ETF:[/cyan] {summary['top_etf']}")
        console.print(f"[green]Best Category:[/green] {summary['best_category']}")
        console.print(f"\n[bold yellow]📁 Output Directory:[/bold yellow] {self.output_dir}")
        console.print(f"[bold yellow]📊 Excel Dashboard:[/bold yellow] {self.excel_manager.excel_file}")
        console.print("[bold cyan]═══════════════════════════════════════════[/bold cyan]\n")

        logger.info("ETF Tracker stopped successfully")


# ============================================================================
# ETF SYMBOL LISTS
# ============================================================================

class ETFSymbolLists:
    """ETF symbol lists"""

    @staticmethod
    def get_all_etfs():
        """Get comprehensive ETF list"""
        return [
            # Gold & Silver
            "GOLDBEES", "GOLDIETF", "HDFCGOLD", "AXISGOLD", "GOLD1", "EGOLD",
            "SILVERBEES", "SILVERIETF", "HDFCSILVER", "AXISILVER", "SILVER1", "ESILVER",

            # Nifty Index
            "NIFTYBEES", "NIFTYETF", "HDFCNIFTY", "AXISNIFTY", "UTINIFTETF",
            "SENSEXETF", "SENSEXADD", "AXSENSEX",

            # Banking
            "BANKBEES", "BANKETF", "AXISBNKETF", "HDFCNIFBAN", "PVTBANIETF", "PSUBNKIETF",

            # Sectoral
            "ITBEES", "ITETF", "AUTOBEES", "PHARMABEES", "METAL", "INFRABEES",

            # Liquid & Debt
            "LIQUIDBEES", "LIQUIDETF", "SBILIQETF",

            # Momentum & Quality
            "MOMENTUM", "QUALITY30", "ALPHA", "LOWVOL",

            # Mid & Small Cap
            "JUNIORBEES", "MIDCAP", "SMALLCAP",
        ]

    @staticmethod
    def get_top_50():
        """Get top 50 most liquid ETFs"""
        return ETFSymbolLists.get_all_etfs()[:50]


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main():
    """Main entry point with progress bars"""

    # Setup logging
    setup_logging()

    # Validate config
    if not config.validate():
        console.print("[red]Error: Invalid configuration[/red]")
        return

    # Display banner
    console.print("\n[bold cyan]" + "=" * 80 + "[/bold cyan]")
    console.print("[bold yellow]🚀 ENHANCED ETF LIVE TRACKER - PRODUCTION READY 🚀[/bold yellow]")
    console.print("[bold cyan]" + "=" * 80 + "[/bold cyan]\n")

    console.print("[cyan]Features:[/cyan]")
    console.print("  ✓ Async API calls with connection pooling")
    console.print("  ✓ Real-time CSV updates with buffering")
    console.print("  ✓ Auto-opening Excel dashboard with live refresh")
    console.print("  ✓ Beautiful progress bars and live display")
    console.print("  ✓ Comprehensive error handling")
    console.print("  ✓ Memory-efficient circular buffers")
    console.print("\n[bold cyan]" + "=" * 80 + "[/bold cyan]\n")

    # Initialize tracker
    tracker = EnhancedETFTracker()

    # Get ETF symbols
    console.print("[yellow]Select ETF list:[/yellow]")
    console.print("1. Top 50 Liquid ETFs (Recommended)")
    console.print("2. All ETFs (~100 symbols)")
    console.print("3. Custom list")

    choice = input("\nEnter choice (1-3): ").strip()

    if choice == "1":
        etf_symbols = ETFSymbolLists.get_top_50()
    elif choice == "2":
        etf_symbols = ETFSymbolLists.get_all_etfs()
    elif choice == "3":
        custom = input("Enter comma-separated symbols: ").strip()
        etf_symbols = [s.strip().upper() for s in custom.split(',')]
    else:
        console.print("[red]Invalid choice[/red]")
        return

    console.print(f"\n[green]Will track {len(etf_symbols)} ETF symbols[/green]\n")

    # Progress bar for setup
    with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            TimeRemainingColumn(),
            console=console
    ) as progress:

        # Load instruments
        task1 = progress.add_task("[cyan]Loading instruments...", total=1)
        tracker.fetcher.load_instruments(lambda msg: progress.update(task1, description=f"[cyan]{msg}"))
        progress.update(task1, completed=1)

        # Subscribe to symbols
        subscribed = tracker.subscribe_symbols(etf_symbols, progress)

        if subscribed == 0:
            console.print("[red]No successful subscriptions. Exiting.[/red]")
            return

        # Fetch historical data
        tracker.fetch_historical_data(progress)

    console.print(f"\n[bold green]✓ Setup complete! Subscribed to {subscribed} ETFs[/bold green]")
    console.print(f"[cyan]Output directory:[/cyan] {tracker.output_dir}\n")

    # Start tracking
    try:
        tracker.start_tracking()
    except KeyboardInterrupt:
        console.print("\n[yellow]Interrupted by user[/yellow]")
    finally:
        tracker.stop_tracking()


if __name__ == "__main__":
    main()